
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def save_to_excel(images):
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "File name"
    ws["B1"] = "Passed/Failed"

    c = 1
    for image in images:
        c += 1
        ws[f"A{c}"] = image["name"]
        ws[f"B{c}"] = image["status"]

        if image["status"] == "passed":
            ws[f"B{c}"].fill = PatternFill("solid", start_color="00FF00")
        else:
            ws[f"B{c}"].fill = PatternFill("solid", start_color="FF0000")

    wb.save("spreadsheets/anonymisation_results.xlsx")